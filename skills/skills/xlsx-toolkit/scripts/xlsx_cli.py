#!/usr/bin/env python3
"""xlsx-toolkit CLI — read/summary/to-csv/to-md subcommands.

MVP scope (read-only):
- read:    cell values → markdown table (all sheets or one)
- summary: sheet list + row/col counts + first-row header preview
- to-csv:  CSV output (stdout or file)
- to-md:   markdown table output

Uses openpyxl 3.1.5 from _sys/skills/env/.venv.
"""

from __future__ import annotations

import argparse
import csv
import io
import sys
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("[xlsx-toolkit] openpyxl 미설치. _sys/skills/env/.venv 확인", file=sys.stderr)
    sys.exit(2)


LARGE_FILE_THRESHOLD_MB = 50
LARGE_FILE_DEFAULT_MAX_ROWS = 100


# ──────────────────────────────────────────────────────
# Loading
# ──────────────────────────────────────────────────────
def load_workbook(path: Path, values_only: bool, read_only: Optional[bool] = None):
    """Load workbook with auto read_only for large files."""
    if read_only is None:
        size_mb = path.stat().st_size / (1024 * 1024)
        read_only = size_mb > LARGE_FILE_THRESHOLD_MB
        if read_only:
            print(
                f"[xlsx-toolkit] 파일 크기 {size_mb:.1f}MB > {LARGE_FILE_THRESHOLD_MB}MB: read_only 모드",
                file=sys.stderr,
            )
    try:
        wb = openpyxl.load_workbook(
            filename=str(path),
            read_only=read_only,
            data_only=values_only,
            keep_vba=False,
        )
    except Exception as e:
        print(f"[xlsx-toolkit] openpyxl load 실패: {e}", file=sys.stderr)
        sys.exit(2)
    return wb, read_only


def resolve_sheet(wb, sheet_ref: Optional[str]) -> Worksheet:
    """Resolve sheet by 1-based index, name, or default to active."""
    if sheet_ref is None:
        return wb.active
    # Try as int (1-based)
    try:
        idx = int(sheet_ref) - 1
        if 0 <= idx < len(wb.sheetnames):
            return wb[wb.sheetnames[idx]]
        print(
            f"[xlsx-toolkit] 시트 #{sheet_ref} 없음 (1..{len(wb.sheetnames)})",
            file=sys.stderr,
        )
        sys.exit(3)
    except ValueError:
        pass
    # Try as name
    if sheet_ref in wb.sheetnames:
        return wb[sheet_ref]
    print(
        f"[xlsx-toolkit] 시트 '{sheet_ref}' 없음. 가용: {wb.sheetnames}",
        file=sys.stderr,
    )
    sys.exit(3)


# ──────────────────────────────────────────────────────
# Cell extraction
# ──────────────────────────────────────────────────────
def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _expand_merged_cells(ws: Worksheet) -> dict:
    """Return {(row, col): value} for cells that should inherit a merged-range value."""
    if getattr(ws, "read_only", False):
        return {}  # read_only mode cannot access merged_cells
    inherited = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                if (r, c) != (merged_range.min_row, merged_range.min_col):
                    inherited[(r, c)] = top_left
    return inherited


def extract_rows(
    ws: Worksheet, max_rows: Optional[int] = None
) -> Tuple[List[List[str]], int, int]:
    """Extract rows as list of list-of-str. Returns (rows, total_rows_seen, total_cols)."""
    merged = _expand_merged_cells(ws)
    rows: List[List[str]] = []
    total_rows = 0
    max_cols = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        total_rows += 1
        if max_rows is not None and len(rows) >= max_rows:
            # keep scanning to count total_rows, but cheaply — for read_only mode break
            if getattr(ws, "read_only", False):
                break
            continue
        str_row = []
        for cell_idx, cell in enumerate(row, start=1):
            val = cell.value
            if val is None and (row_idx, cell_idx) in merged:
                val = merged[(row_idx, cell_idx)]
            str_row.append(_cell_str(val))
        # Trim trailing empty cells
        while str_row and str_row[-1] == "":
            str_row.pop()
        if str_row:  # skip fully empty rows
            max_cols = max(max_cols, len(str_row))
            rows.append(str_row)

    # Pad rows to max_cols
    for r in rows:
        while len(r) < max_cols:
            r.append("")

    return rows, total_rows, max_cols


# ──────────────────────────────────────────────────────
# Output formatters
# ──────────────────────────────────────────────────────
def rows_to_markdown(rows: List[List[str]], sheet_name: str) -> str:
    if not rows:
        return f"### {sheet_name}\n\n_(empty)_\n"
    out = [f"### {sheet_name}", ""]
    header = rows[0]
    body = rows[1:] if len(rows) > 1 else []
    out.append("| " + " | ".join(_md_escape(c) for c in header) + " |")
    out.append("| " + " | ".join("---" for _ in header) + " |")
    for r in body:
        out.append("| " + " | ".join(_md_escape(c) for c in r) + " |")
    out.append("")
    return "\n".join(out)


def _md_escape(s: str) -> str:
    return s.replace("|", r"\|").replace("\n", " ")


def rows_to_csv(rows: List[List[str]]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return buf.getvalue()


# ──────────────────────────────────────────────────────
# Subcommands
# ──────────────────────────────────────────────────────
def cmd_read(args: argparse.Namespace) -> int:
    wb, _ = load_workbook(Path(args.file), args.values_only)
    if args.sheet is None:
        # All sheets
        out_parts = [f"# {Path(args.file).name}", ""]
        for name in wb.sheetnames:
            ws = wb[name]
            rows, total, cols = extract_rows(ws, args.max_rows)
            out_parts.append(rows_to_markdown(rows, name))
            if args.max_rows and total > args.max_rows:
                out_parts.append(f"_... ({total - args.max_rows} more rows truncated)_\n")
        sys.stdout.write("\n".join(out_parts))
    else:
        ws = resolve_sheet(wb, args.sheet)
        rows, total, cols = extract_rows(ws, args.max_rows)
        sys.stdout.write(rows_to_markdown(rows, ws.title))
        if args.max_rows and total > args.max_rows:
            sys.stdout.write(f"\n_... ({total - args.max_rows} more rows truncated)_\n")
    return 0


def cmd_summary(args: argparse.Namespace) -> int:
    path = Path(args.file)
    wb, read_only = load_workbook(path, values_only=False)
    size_mb = path.stat().st_size / (1024 * 1024)
    print(f"# {path.name}")
    print(f"- Size: {size_mb:.2f} MB")
    print(f"- Sheets: {len(wb.sheetnames)}")
    if read_only:
        print("- Mode: read_only (large file)")
    print()
    print("| # | Sheet | Rows | Cols | Header |")
    print("|---|---|---|---|---|")
    for i, name in enumerate(wb.sheetnames, start=1):
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        # First row as header preview
        first_row = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            first_row = [_cell_str(v) for v in row if v is not None]
            break
        header_preview = ", ".join(first_row[:6])
        if len(first_row) > 6:
            header_preview += ", ..."
        print(
            f"| {i} | {_md_escape(name)} | {max_row} | {max_col} | {_md_escape(header_preview)} |"
        )
    return 0


def cmd_to_csv(args: argparse.Namespace) -> int:
    wb, _ = load_workbook(Path(args.file), args.values_only)

    def emit(rows: List[List[str]], target):
        target.write(rows_to_csv(rows))

    if args.all_sheets:
        base = Path(args.output or Path(args.file).stem).with_suffix("")
        for name in wb.sheetnames:
            ws = wb[name]
            rows, _, _ = extract_rows(ws, args.max_rows)
            out_path = base.with_name(f"{base.name}-{_sanitize(name)}.csv")
            with out_path.open("w", encoding="utf-8", newline="") as f:
                emit(rows, f)
            print(f"wrote {out_path} ({len(rows)} rows)", file=sys.stderr)
        return 0

    ws = resolve_sheet(wb, args.sheet)
    rows, _, _ = extract_rows(ws, args.max_rows)
    if args.output:
        with Path(args.output).open("w", encoding="utf-8", newline="") as f:
            emit(rows, f)
        print(f"wrote {args.output} ({len(rows)} rows)", file=sys.stderr)
    else:
        emit(rows, sys.stdout)
    return 0


def cmd_to_md(args: argparse.Namespace) -> int:
    wb, _ = load_workbook(Path(args.file), args.values_only)
    if args.sheet is None:
        out_parts = [f"# {Path(args.file).name}", ""]
        for name in wb.sheetnames:
            ws = wb[name]
            rows, _, _ = extract_rows(ws, args.max_rows)
            out_parts.append(rows_to_markdown(rows, name))
        content = "\n".join(out_parts)
    else:
        ws = resolve_sheet(wb, args.sheet)
        rows, _, _ = extract_rows(ws, args.max_rows)
        content = rows_to_markdown(rows, ws.title)
    if args.output:
        Path(args.output).write_text(content, encoding="utf-8")
        print(f"wrote {args.output}", file=sys.stderr)
    else:
        sys.stdout.write(content)
    return 0


def _sanitize(name: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in name)


# ──────────────────────────────────────────────────────
# Argparse
# ──────────────────────────────────────────────────────
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="xlsx", description="xlsx-toolkit — Excel read-only CLI"
    )
    sub = p.add_subparsers(dest="command", required=True)

    def _common(sp):
        sp.add_argument("file", help=".xlsx file path")
        sp.add_argument("--sheet", help="sheet index (1-based) or name")
        sp.add_argument("--max-rows", type=int, help="limit rows per sheet")
        sp.add_argument(
            "--values-only",
            action="store_true",
            help="use cached values instead of formulas",
        )

    sp_read = sub.add_parser("read", help="read sheets → markdown")
    _common(sp_read)
    sp_read.set_defaults(func=cmd_read)

    sp_sum = sub.add_parser("summary", help="workbook summary")
    sp_sum.add_argument("file", help=".xlsx file path")
    sp_sum.set_defaults(func=cmd_summary)

    sp_csv = sub.add_parser("to-csv", help="export as CSV")
    _common(sp_csv)
    sp_csv.add_argument("-o", "--output", help="output file (default stdout)")
    sp_csv.add_argument(
        "--all-sheets",
        action="store_true",
        help="write one CSV per sheet (output is base name)",
    )
    sp_csv.set_defaults(func=cmd_to_csv)

    sp_md = sub.add_parser("to-md", help="export as markdown table")
    _common(sp_md)
    sp_md.add_argument("-o", "--output", help="output file (default stdout)")
    sp_md.set_defaults(func=cmd_to_md)

    return p


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if not Path(args.file).exists():
        print(f"[xlsx-toolkit] 파일 없음: {args.file}", file=sys.stderr)
        return 1
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
